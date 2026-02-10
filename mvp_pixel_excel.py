import argparse
import sys
from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def load_and_process_image(input_path, target_width, target_height, threshold, invert):
    try:
        img = Image.open(input_path).convert('L')
        
        if target_height is None:
            aspect_ratio = img.height / img.width
            target_height = int(target_width * aspect_ratio)
        
        img = img.resize((target_width, target_height), Image.Resampling.LANCZOS)
        
        pixels = []
        for y in range(target_height):
            row = []
            for x in range(target_width):
                pixel_value = img.getpixel((x, y))
                is_black = pixel_value < threshold
                if invert:
                    is_black = not is_black
                row.append(is_black)
            pixels.append(row)
        
        return pixels, target_width, target_height
    
    except FileNotFoundError:
        print(f"Error: Input file '{input_path}' not found.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error processing image: {e}", file=sys.stderr)
        sys.exit(1)


def create_pixel_art_excel(pixels, output_path, width, height):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pixel Art"
    
    BLACK_CHAR = "█"  
    WHITE_CHAR = "░"  
    
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row_idx, row in enumerate(pixels, start=1):
        for col_idx, is_black in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = BLACK_CHAR if is_black else WHITE_CHAR
            cell.fill = black_fill if is_black else white_fill
    
    cell_size_width = 2.14  
    cell_size_height = 15   
    
    for col_idx in range(1, width + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = cell_size_width
    
    for row_idx in range(1, height + 1):
        ws.row_dimensions[row_idx].height = cell_size_height
    
    ws.sheet_view.showGridLines = False
    
    try:
        wb.save(output_path)
    except Exception as e:
        print(f"Error saving Excel file: {e}", file=sys.stderr)
        sys.exit(1)


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Convert an image to black-and-white pixel art in Excel"
    )
    
    parser.add_argument(
        "--input",
        required=True,
    )
    
    parser.add_argument(
        "--output",
        required=True,
    )
    
    parser.add_argument(
        "--width",
        type=int,
        required=True,
    )
    
    parser.add_argument(
        "--height",
        type=int,
        default=None,
    )
    
    parser.add_argument(
        "--threshold",
        type=int,
        default=128,
    )
    
    parser.add_argument(
        "--invert",
        action="store_true",
    )
    
    return parser.parse_args()


def main():
    """Main entry point."""
    args = parse_arguments()
    
    if args.width <= 0:
        print("Error: Width must be positive.", file=sys.stderr)
        sys.exit(1)
    
    if args.height is not None and args.height <= 0:
        print("Error: Height must be positive.", file=sys.stderr)
        sys.exit(1)
    
    if not (0 <= args.threshold <= 255):
        print("Error: Threshold must be between 0 and 255.", file=sys.stderr)
        sys.exit(1)
    
    print(f"Loading and processing image: {args.input}")
    pixels, final_width, final_height = load_and_process_image(
        args.input,
        args.width,
        args.height,
        args.threshold,
        args.invert
    )
    
    print(f"Creating Excel file: {args.output}")
    create_pixel_art_excel(pixels, args.output, final_width, final_height)
    
    print(f"✓ Success! Created {final_width}x{final_height} pixel art in {args.output}")


if __name__ == "__main__":
    main()
